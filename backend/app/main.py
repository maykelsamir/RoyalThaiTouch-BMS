from datetime import date, datetime, timedelta
import io
import os
import time
from typing import List, Optional

from fastapi import Depends, FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel
from sqlalchemy import Boolean, Column, Date, DateTime, ForeignKey, Integer, Numeric, String, Text, UniqueConstraint, create_engine, func
from sqlalchemy.exc import OperationalError
from sqlalchemy.orm import Session, declarative_base, sessionmaker

DATABASE_URL = os.getenv("DATABASE_URL", "postgresql://postgres:postgres@localhost:5432/royalthaitouch")
engine = create_engine(DATABASE_URL)
SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
Base = declarative_base()
app = FastAPI(title="Royal Thai Touch ERP", version="0.8.6")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

class Branch(Base):
    __tablename__ = "branches"
    id = Column(Integer, primary_key=True)
    name = Column(String(120), unique=True, nullable=False)
    address = Column(String(255), nullable=True)
    active = Column(Boolean, default=True, nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow)

class DailyRevenue(Base):
    __tablename__ = "daily_revenues"
    __table_args__ = (UniqueConstraint("branch_id", "business_date", name="uq_revenue_branch_date"),)
    id = Column(Integer, primary_key=True)
    branch_id = Column(Integer, ForeignKey("branches.id"), nullable=False)
    business_date = Column(Date, nullable=False)
    amount = Column(Numeric(15, 2), nullable=False, default=0)
    notes = Column(Text, nullable=True)
    closed = Column(Boolean, default=False, nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow)

class Expense(Base):
    __tablename__ = "expenses"
    id = Column(Integer, primary_key=True)
    branch_id = Column(Integer, ForeignKey("branches.id"), nullable=False)
    business_date = Column(Date, nullable=False)
    category = Column(String(120), nullable=False, default="Other Expenses")
    amount = Column(Numeric(15, 2), nullable=False, default=0)
    notes = Column(Text, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)

class AuditLog(Base):
    __tablename__ = "audit_logs"
    id = Column(Integer, primary_key=True)
    action = Column(String(120), nullable=False)
    entity = Column(String(120), nullable=False)
    details = Column(Text, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)

class DailyEntryExpense(BaseModel):
    category: str = "Other Expenses"
    amount: float
    notes: Optional[str] = None

class DailyEntryInput(BaseModel):
    branch_id: int
    business_date: date
    revenue: float = 0
    notes: Optional[str] = None
    expenses: List[DailyEntryExpense] = []

class BranchInput(BaseModel):
    name: str
    address: Optional[str] = "Erbil"
    active: bool = True

def wait_for_database() -> None:
    for _ in range(40):
        try:
            Base.metadata.create_all(bind=engine)
            return
        except OperationalError:
            time.sleep(2)
    Base.metadata.create_all(bind=engine)

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def seed_data(db: Session) -> None:
    for name in ["Canyon Thai Spa", "Sheratoon Thai Spa", "Hyksos Thai Spa", "Crixus Thai Spa"]:
        if not db.query(Branch).filter(Branch.name == name).first():
            db.add(Branch(name=name, address="Erbil"))
    db.commit()

def create_audit(db: Session, action: str, entity: str, details: str = "") -> None:
    db.add(AuditLog(action=action, entity=entity, details=details))

def active_branches(db: Session, branch_id: Optional[int] = None):
    q = db.query(Branch).filter(Branch.active == True)
    if branch_id:
        q = q.filter(Branch.id == branch_id)
    return q.order_by(Branch.id).all()

def branch_summary(db: Session, target_date: date, branch_id: Optional[int] = None, include_expenses: bool = True):
    branches = active_branches(db, branch_id)
    revenues = db.query(DailyRevenue).filter(DailyRevenue.business_date == target_date).all()
    expenses = db.query(Expense).filter(Expense.business_date == target_date).all() if include_expenses else []
    rows = []
    for branch in branches:
        rev = next((r for r in revenues if r.branch_id == branch.id), None)
        branch_expenses = [e for e in expenses if e.branch_id == branch.id]
        total_expenses = sum(float(e.amount) for e in branch_expenses) if include_expenses else 0
        revenue_amount = float(rev.amount) if rev else 0
        rows.append({"branch_id": branch.id, "branch": branch.name, "revenue": revenue_amount, "expenses": total_expenses, "net_profit": revenue_amount - total_expenses, "status": "Closed" if rev and rev.closed else ("Submitted" if rev else "Missing"), "expense_count": len(branch_expenses)})
    return rows

def report_data(db: Session, date_from: date, date_to: date, branch_id: Optional[int] = None, include_expenses: bool = True):
    current = date_from
    daily_rows = []
    branch_totals = {}
    while current <= date_to:
        rows = branch_summary(db, current, branch_id, include_expenses)
        day_revenue = sum(r["revenue"] for r in rows)
        day_expenses = sum(r["expenses"] for r in rows) if include_expenses else 0
        daily_rows.append({"date": str(current), "revenue": day_revenue, "expenses": day_expenses, "net_profit": day_revenue - day_expenses})
        for row in rows:
            key = row["branch"]
            if key not in branch_totals:
                branch_totals[key] = {"branch": key, "revenue": 0, "expenses": 0, "net_profit": 0}
            branch_totals[key]["revenue"] += row["revenue"]
            branch_totals[key]["expenses"] += row["expenses"] if include_expenses else 0
            branch_totals[key]["net_profit"] += row["net_profit"]
        current += timedelta(days=1)
    total_revenue = sum(x["revenue"] for x in daily_rows)
    total_expenses = sum(x["expenses"] for x in daily_rows) if include_expenses else 0
    branch_name = "All Branches"
    if branch_id:
        branch = db.query(Branch).filter(Branch.id == branch_id).first()
        branch_name = branch.name if branch else f"Branch {branch_id}"
    return {"date_from": str(date_from), "date_to": str(date_to), "branch_id": branch_id, "branch_name": branch_name, "include_expenses": include_expenses, "total_revenue": total_revenue, "total_expenses": total_expenses, "net_profit": total_revenue - total_expenses, "daily_rows": daily_rows, "branch_totals": list(branch_totals.values())}

@app.on_event("startup")
def startup():
    wait_for_database()
    db = SessionLocal()
    try:
        seed_data(db)
    finally:
        db.close()

@app.get("/")
def root():
    return {"application": "Royal Thai Touch ERP", "status": "running", "version": "0.8.6"}

@app.get("/health")
def health():
    return {"status": "healthy", "app": "Royal Thai Touch ERP", "version": "0.8.6"}

@app.get("/branches")
def get_branches(db: Session = Depends(get_db)):
    return active_branches(db)

@app.post("/branches")
def create_branch(body: BranchInput, db: Session = Depends(get_db)):
    if db.query(Branch).filter(Branch.name == body.name).first():
        raise HTTPException(status_code=400, detail="Branch already exists")
    branch = Branch(name=body.name, address=body.address, active=body.active)
    db.add(branch)
    create_audit(db, "CREATE", "Branch", body.name)
    db.commit()
    db.refresh(branch)
    return branch

@app.get("/dashboard")
def dashboard(business_date: Optional[date] = Query(None), db: Session = Depends(get_db)):
    target_date = business_date or date.today()
    rows = branch_summary(db, target_date)
    total_revenue = sum(r["revenue"] for r in rows)
    total_expenses = sum(r["expenses"] for r in rows)
    month_start = target_date.replace(day=1)
    month_revenue = db.query(func.coalesce(func.sum(DailyRevenue.amount), 0)).filter(DailyRevenue.business_date >= month_start, DailyRevenue.business_date <= target_date).scalar()
    month_expenses = db.query(func.coalesce(func.sum(Expense.amount), 0)).filter(Expense.business_date >= month_start, Expense.business_date <= target_date).scalar()
    best_branch = max(rows, key=lambda r: r["net_profit"], default=None)
    return {"date": str(target_date), "total_revenue": total_revenue, "total_expenses": total_expenses, "net_profit": total_revenue - total_expenses, "month_revenue": float(month_revenue or 0), "month_expenses": float(month_expenses or 0), "month_profit": float(month_revenue or 0) - float(month_expenses or 0), "best_branch": best_branch["branch"] if best_branch and best_branch["net_profit"] > 0 else "-", "missing_branches": [r["branch"] for r in rows if r["status"] == "Missing"], "branches": rows}

@app.get("/dashboard/trend")
def dashboard_trend(days: int = 30, db: Session = Depends(get_db)):
    end = date.today()
    start = end - timedelta(days=max(1, min(days, 365)) - 1)
    rows = []
    current = start
    while current <= end:
        summary = branch_summary(db, current)
        revenue = sum(x["revenue"] for x in summary)
        expenses = sum(x["expenses"] for x in summary)
        rows.append({"date": str(current), "revenue": revenue, "expenses": expenses, "profit": revenue - expenses})
        current += timedelta(days=1)
    return rows

@app.post("/daily-entry")
def save_daily_entry(body: DailyEntryInput, db: Session = Depends(get_db)):
    if body.revenue < 0:
        raise HTTPException(status_code=400, detail="Revenue cannot be negative")
    for item in body.expenses:
        if item.amount < 0:
            raise HTTPException(status_code=400, detail="Expense amount cannot be negative")
    branch = db.query(Branch).filter(Branch.id == body.branch_id, Branch.active == True).first()
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")
    revenue = db.query(DailyRevenue).filter(DailyRevenue.branch_id == body.branch_id, DailyRevenue.business_date == body.business_date).first()
    if revenue and revenue.closed:
        raise HTTPException(status_code=403, detail="Day is closed")
    if not revenue:
        revenue = DailyRevenue(branch_id=body.branch_id, business_date=body.business_date, amount=body.revenue, notes=body.notes)
        db.add(revenue)
        action = "CREATE_DAILY_ENTRY"
    else:
        revenue.amount = body.revenue
        revenue.notes = body.notes
        action = "UPDATE_DAILY_ENTRY"
    db.query(Expense).filter(Expense.branch_id == body.branch_id, Expense.business_date == body.business_date).delete()
    for item in body.expenses:
        if item.amount > 0:
            db.add(Expense(branch_id=body.branch_id, business_date=body.business_date, category=item.category or "Other Expenses", amount=item.amount, notes=item.notes))
    create_audit(db, action, "DailyEntry", f"branch={branch.name}; date={body.business_date}")
    db.commit()
    return {"status": "saved"}

@app.get("/daily-entry")
def get_daily_entry(branch_id: int, business_date: date, db: Session = Depends(get_db)):
    revenue = db.query(DailyRevenue).filter(DailyRevenue.branch_id == branch_id, DailyRevenue.business_date == business_date).first()
    expenses = db.query(Expense).filter(Expense.branch_id == branch_id, Expense.business_date == business_date).order_by(Expense.id).all()
    return {"branch_id": branch_id, "business_date": str(business_date), "revenue": float(revenue.amount) if revenue else 0, "notes": revenue.notes if revenue else "", "closed": bool(revenue.closed) if revenue else False, "expenses": [{"id": e.id, "category": e.category, "amount": float(e.amount), "notes": e.notes or ""} for e in expenses]}

@app.post("/daily-entry/close")
def close_day(branch_id: int, business_date: date, db: Session = Depends(get_db)):
    revenue = db.query(DailyRevenue).filter(DailyRevenue.branch_id == branch_id, DailyRevenue.business_date == business_date).first()
    if not revenue:
        raise HTTPException(status_code=404, detail="No revenue entry found")
    revenue.closed = True
    create_audit(db, "CLOSE_DAY", "DailyEntry", f"branch_id={branch_id}; date={business_date}")
    db.commit()
    return {"status": "closed"}

@app.get("/reports/summary")
def report_summary(date_from: date, date_to: date, branch_id: Optional[int] = None, include_expenses: bool = True, db: Session = Depends(get_db)):
    return report_data(db, date_from, date_to, branch_id, include_expenses)

@app.get("/reports/excel")
def export_excel(date_from: date, date_to: date, branch_id: Optional[int] = None, include_expenses: bool = True, db: Session = Depends(get_db)):
    data = report_data(db, date_from, date_to, branch_id, include_expenses)
    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Report"
    gold, black, white = "D4AF37", "111111", "FFFFFF"
    ws.merge_cells("A1:E1")
    ws["A1"] = "Royal Thai Touch ERP - Financial Report"
    ws["A1"].font = Font(bold=True, size=16, color=gold)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = PatternFill("solid", fgColor=black)
    ws.merge_cells("A2:E2")
    ws["A2"] = f"Branch: {data['branch_name']} | Period: {date_from} to {date_to} | Expenses: {'Included' if include_expenses else 'Excluded'}"
    ws["A2"].alignment = Alignment(horizontal="center")
    headers = ["Date", "Revenue"] + (["Expenses"] if include_expenses else []) + ["Net Profit"]
    ws.append([])
    ws.append(headers)
    for cell in ws[4]:
        cell.font = Font(bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=black)
        cell.border = Border(bottom=Side(style="thin", color="999999"))
    for row in data["daily_rows"]:
        values = [row["date"], row["revenue"]]
        if include_expenses:
            values.append(row["expenses"])
        values.append(row["net_profit"])
        ws.append(values)
    last_row = ws.max_row + 2
    ws[f"A{last_row}"] = "TOTAL"
    ws[f"B{last_row}"] = data["total_revenue"]
    col_profit = "D" if include_expenses else "C"
    if include_expenses:
        ws[f"C{last_row}"] = data["total_expenses"]
    ws[f"{col_profit}{last_row}"] = data["net_profit"]
    for cell in ws[last_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F4E4A6")
    for col in ["A", "B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 24
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    suffix = "with-expenses" if include_expenses else "revenue-profit-only"
    filename = f"royal-thai-touch-{data['branch_name'].replace(' ', '-')}-{date_from}-to-{date_to}-{suffix}.xlsx"
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})

@app.get("/reports/pdf")
def export_pdf(date_from: date, date_to: date, branch_id: Optional[int] = None, include_expenses: bool = True, db: Session = Depends(get_db)):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    data = report_data(db, date_from, date_to, branch_id, include_expenses)
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=28, leftMargin=28, topMargin=28, bottomMargin=28)
    styles = getSampleStyleSheet()
    story = [Paragraph("Royal Thai Touch ERP", styles["Title"]), Paragraph(f"Monthly Report - {data['branch_name']}", styles["Heading2"]), Paragraph(f"Period: {date_from} to {date_to}", styles["Normal"]), Paragraph(f"Expenses: {'Included' if include_expenses else 'Excluded - Revenue and profit only'}", styles["Normal"]), Spacer(1, 12), Paragraph(f"Total Revenue: IQD {data['total_revenue']:,.0f}", styles["Normal"])]
    if include_expenses:
        story.append(Paragraph(f"Total Expenses: IQD {data['total_expenses']:,.0f}", styles["Normal"]))
    story.append(Paragraph(f"Net Profit: IQD {data['net_profit']:,.0f}", styles["Normal"]))
    story.append(Spacer(1, 14))
    table_data = [["Date", "Revenue"] + (["Expenses"] if include_expenses else []) + ["Net Profit"]]
    for row in data["daily_rows"]:
        values = [row["date"], f"IQD {row['revenue']:,.0f}"]
        if include_expenses:
            values.append(f"IQD {row['expenses']:,.0f}")
        values.append(f"IQD {row['net_profit']:,.0f}")
        table_data.append(values)
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111111")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#D4AF37")), ("GRID", (0, 0), (-1, -1), 0.5, colors.grey), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("ALIGN", (1, 1), (-1, -1), "RIGHT"), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F7F7")])]))
    story.append(table)
    story.append(Spacer(1, 16))
    story.append(Paragraph("Confidential - Royal Thai Touch", styles["Italic"]))
    doc.build(story)
    output.seek(0)
    suffix = "with-expenses" if include_expenses else "revenue-profit-only"
    filename = f"royal-thai-touch-{data['branch_name'].replace(' ', '-')}-{date_from}-to-{date_to}-{suffix}.pdf"
    return StreamingResponse(output, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})

@app.get("/audit-logs")
def audit_logs(db: Session = Depends(get_db)):
    return db.query(AuditLog).order_by(AuditLog.id.desc()).limit(100).all()
